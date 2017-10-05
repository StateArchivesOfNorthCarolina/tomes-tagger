#!/usr/bin/env python3

"""
This module ... ???

Todo:
    * Should you validate the data type for each row, i.e. "_validate_row()"?
        - It won't add to the length of stanfordize_file().
        - No: Just trust that data is OK unless we start to see data entry errors. :-]
    * You need to catch SameFileError when copying text to text OR ... should you only
    support Excel files (yes?!).
"""

# import modules.
import csv
import glob
import logging
import os
import shutil
from openpyxl import load_workbook


class XLSXToStanford():
    """ A class for ... ??? """


    def __init__(self, entity_worksheet="Entities", charset="utf-8"):
        """ Sets instance attributes. """
        
        # set logger; suppress logging by default. 
        self.logger = logging.getLogger(__name__)
        self.logger.addHandler(logging.NullHandler())

        # ???
        self.entity_worksheet = entity_worksheet
        
        # assume column number for required headers.
        self.required_headers = ("pattern", "description", "case_sensitive", "label", 
                "authority")


    def _get_pattern(self, pattern, is_case_sensitive):
        """ ??? """

        # ???
        self.logger.debug("Inspecting pattern: {}".format(pattern))

        # remove excess whitespace from @pattern.
        _pattern = " ".join(pattern.strip().split())
        if pattern != _pattern:
            self.logger.warning("Removing excess whitespace in: {}".format(pattern))
            pattern = _pattern
        
        # alter @pattern to be case-insensitive if @is_case_sensitive is True.
        if is_case_sensitive:
            self.logger.debug("Rewriting case-insensitive pattern to: {}".format(pattern))
            pattern = "(?i)" + pattern.replace(" ", " (?i)")
        
        return pattern


    def _validate_header_row(self, header_row):
        """ ??? """
        
        # ???
        self.logger.debug("Validating header row: {}".format(header_row))
        
        # ???
        if header_row == self.required_headers:
            self.logger.debug("Header row is valid.")
            return True
        
        # ???
        missing_headers = [header for header in self.required_headers if header not in
                header_row]
        if len(missing_headers) != 0:
            missing_headers = ", ".join(missing_headers)
            self.logger.error("Header row is not valid. Missing: {}".format(missing_headers))
  
        return False


    def _get_tsv_file(self, xlsx_file, tsv_file=None):
        """ ??? """
        
        # ???
        if tsv_file is None:
            tsv_file = xlsx_file.replace("__mapping.xlsx", "__mapping.txt")
        if os.path.isfile(tsv_file):
            self.logger.warning("File '{}' already exists.".format(
                tsv_file))

        return tsv_file

        
    def _get_entity_rows(self, xlsx_file):
        """ ??? """

        # load workbook; verify that required worksheet exists.
        workbook = load_workbook(xlsx_file, read_only=False, data_only=True)
        try:
            entity_rows = workbook[self.entity_worksheet]
        except KeyError:
            self.logger.error("Missing required worksheet '{}' in workbook: {}".format(
                    self.entity_worksheet, xlsx_file))
            raise

        return entity_rows


    def stanfordize_file(self, xlsx_file, tsv_file=None):
        """ ??? """
            
        # ???
        self.logger.info("Loading workbook: {}".format(xlsx_file))
        entity_rows = self._get_entity_rows(xlsx_file)
        tsv_file = self._get_tsv_file(xlsx_file, tsv_file)

        # ???
        tsv = open(tsv_file, "w")
        tsv_writer = csv.writer(tsv, delimiter="\t", lineterminator="")
        
        #
        i = 0
        for row in entity_rows.values:

            i += 1

            # validate header row.
            if i == 1:
                if not self._validate_header_row(row):
                    self.logger.info("Removing file: {}".format(tsv_file))
                    tsv.close()
                    os.remove(tsv_file)
                    err = "Bad header in workbook: {}".format(xlsx_file)
                    raise Exception(err)
                else:
                    self.logger.info("Writing to mapping file: {}".format(tsv_file))
            
            # check for incomplete rows.               
            if None in row:
                self.logger.warning("Empty cell in row #{}. Skipping row.".format(i))
                continue

            # ???
            pattern, description, case_sensitive, label, authority = row
            pattern = self._get_pattern(pattern, case_sensitive)
            label = authority + "/" + label

            if pattern == "":
                self.logger.warning("Illegal pattern in row #{}. Skipping row.".format(i))
                continue

            # write row to file and avoid final linebreak (otherwise CoreNLP will crash).
            if i != 1:
                tsv_writer.writerow([pattern,label])
                if i != entity_rows.max_row:
                    tsv.write("\n")

        return


    def stanfordize_folder(self, input_dir, output_dir):
        """ ??? """
        
        # ???
        input_dir = os.path.abspath(input_dir)
        output_dir = os.path.abspath(output_dir)

        # ???
        for _dir in [input_dir, output_dir]:
            if not os.path.isdir(_dir):
                self.logger.warning("Non-existant directory: {}".format(_dir))
                return

        # glob mapping files; avoid temporary Excel files.
        xlsx_files = glob.glob(input_dir + "/[!~]*__mapping.xlsx")
        txt_files = glob.glob(input_dir + "/*__mapping.txt")

        # ???
        for txt_file in txt_files:
            out_file = txt_file.replace(input_dir, output_dir)
            self.logger.info("Copying '{}' to {}.".format(txt_file, out_file))
            shutil.copy(txt_file, out_file)

        # ???
        for xlsx_file in xlsx_files:
            out_file = self._get_tsv_file(xlsx_file)
            out_file = out_file.replace(input_dir, output_dir)
            self.logger.info("Mapping '{}' to {}.".format(xlsx_file, out_file))
            self.stanfordize_file(xlsx_file, out_file)

        return


# TEST.
if __name__ == "__main__":

    logging.basicConfig(level=logging.DEBUG)

    xlsx_path = "../tests/sample_files/entity_dictionaries"
    xlsx_file = xlsx_path + "/foo__mapping.xlsx"

    x2s = XLSXToStanford()
    #x2s.stanfordize_file(xlsx_file)
    x2s.stanfordize_folder(xlsx_path, xlsx_path)
