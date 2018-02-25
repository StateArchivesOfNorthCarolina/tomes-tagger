#!/usr/bin/env python3

""" This module contains a class for converting a TOMES Excel 2007+ (.xlsx) entity dictionary
file to a Python generator where each item is a dict with headers as keys and row data as 
values. """

# import modules.
import hashlib
import itertools
import logging
import os
from openpyxl import load_workbook


class XLSXToEntities():
    """ A class for converting a TOMES Excel 2007+ (.xlsx) entity dictionary file to a Python
    generator where each item is a dict with headers as keys and row data as values.
    
    Example:
        >>> x2e = XLSXToEntities()
        >>> entities = x2e.get_entities("entities.xlsx")
        >>> next(entities) # mapped dictionary for first data row in Excel file.
    """


    def __init__(self, entity_worksheet="Entities", charset="utf-8"):
        """ Sets instance attributes.
        
        Args:
            - entity_worksheet (str): The name of the worksheet to read from a given 
            workbook, i.e. an Excel file.
            - charset (str): Encoding used when writing to file.
        """
        
        # set logger; suppress logging by default. 
        self.logger = logging.getLogger(__name__)
        self.logger.addHandler(logging.NullHandler())

        # set attributes.
        self.entity_worksheet = entity_worksheet
        self.charset = charset
        self.required_headers = {"identifier": str, "pattern": str, "description": str,
                "case_sensitive": bool, "label": str, "authority": str}


    def _get_hash_prefix(self, xlsx_file):
        """ Gets the first six characters of the SHA-256 hash of @xlsx_file and 
        adds a trailing hash mark.
        
        Args:
            - xlsx_file (str): The path to the Excel file to load.

        Returns:
            str: The return value.
        """

        # placeholder value.
        hash_prefix = ""

        # get checksum of @xlsx_file; truncate it.
        checksum = hashlib.sha256()
        with open(xlsx_file, "rb") as xf:
            xf_bytes = xf.read()
        checksum.update(xf_bytes)
        hash_prefix = checksum.hexdigest()[:6] + "#"

        return hash_prefix


    def _validate_header(self, header_row):
        """ Determines if @header_row is contains all the fields in @self.required_headers.
        
        Args:
            - header_row (tuple): The row values for the presumed first row of data.

        Returns:
            bool: The return value.
            True if @header_row is valid, otherwise False.
        """

        self.logger.info("Validating header row.")
        
        # assume value.
        is_valid = True

        # check if @header_row is perfect.
        if header_row == tuple(self.required_headers.keys()):
            self.logger.info("Header is valid.")
            return is_valid
        else:
            self.logger.error("Header is invalid.")
            is_valid = False

        # report on any missing header fields.
        missing_headers = [header for header in self.required_headers if header not in
                header_row]
        if len(missing_headers) != 0:
            self.logger.warning("Missing required fields: {}".format(missing_headers))
            
        # report on any duplicate fields.
        duplicate_headers = [header for header in header_row if header_row.count(header) != 1]
        if len(duplicate_headers) != 0:
            self.logger.warning("Found duplicate fields: {}".format(set(duplicate_headers)))

        # report on any extra fields.
        extra_headers = [header for header in header_row if header not in 
                self.required_headers]
        if len(extra_headers) != 0:
            self.logger.warning("Found extra fields: {}".format(extra_headers))
        
        return is_valid


    def _validate_row(self, row, row_number):
        """ Determines if @row contains the correct data types per @self.required_headers. 
        
        Args:
            - row (dict): An item from self.get_entities().
            - row_number (int): The line number of @row within @self.entity_worksheet.
            
        Returns:
            bool: The return value.
            True if @row is valid, otherwise False.
        """

        # assume value.
        is_valid = True

        # test if each field in @row has the correct data type.
        tests = []
        for field, value in row.items():
            value_type, header_type = (type(value).__name__, 
                    self.required_headers[field].__name__)
            test = value_type == header_type
            if not test:
                err = "Field '{}' in row {} not valid; expected '{}', got '{}'.".format(
                        field, row_number, header_type, value_type)
                self.logger.debug(err)
            tests.append(test)

        # if any test failed, set @is_valid to False.
        if False in tests:
            is_valid = False
  
        return is_valid


    def _get_tomes_pattern(self, pattern):
        """ Interprets @pattern as a 'TOMES pattern', allowing for single row notation of more
        complex regex patterns. For more information, see the documentation files. NOTE: this
        uses eval(). """
        
        # remove excess whitespace.
        pattern = pattern.strip()

        # test for incorrect TOMES pattern usage.
        if len(pattern) == 0:
            self.logger.warning("TOMES pattern invalid; falling back to empty output.")
            return []     

        # make sure a trailing comma exists; this prevents itertools.product() from splitting
        # something like "TOMES_PATTERN:{'[A|B]'}" from being split as a simple string: '{' +
        # 'A' + '|', etc.
        if pattern[-1] != ",":
            pattern += ","
        
        # interpret the pattern.
        patterns = []
        try:
            pattern = eval(pattern)
            patterns = [i for i in itertools.product(*pattern)]
            patterns.reverse()
        except (NameError, SyntaxError, TypeError) as err:
            self.logger.error(err)
            self.logger.warning("Invalid TOMES pattern syntax; falling back to empty output.")
            self.logger.debug("TOMES pattern: {}".format(pattern))

        return patterns


    def get_manifestations(self, pattern, case_sensitive, row_number):
        """ Returns manifestations of @pattern.
        
        Args:
            - pattern (str): The "pattern" field value for a given row.
            - case_sensitive (bool): The "case_sensitive" field value for a given row.
            - row_number (int): The line number of @row within @self.entity_worksheet.

        Returns:
            list: The return value.
            The altered version(s) of @pattern.
        """
        
        # remove excess whitespace.
        pattern = pattern.strip()

        # assume values.
        manifestations = []
        is_tomes_pattern = False
        
        # if @pattern is a TOMES pattern instance, alter it per self._get_tomes_pattern().
        tomes_pattern = "TOMES_PATTERN:"
        tomes_pattern_len = len(tomes_pattern)
        if pattern[:tomes_pattern_len] == tomes_pattern:
            self.logger.info("Found TOMES pattern in row {}.".format(row_number))
            is_tomes_pattern = True
            pattern = pattern[tomes_pattern_len:]
            manifestations = self._get_tomes_pattern(pattern)

        # if specified, alter @pattern to ignore case provided @is_tomes_pattern is False.
        if not case_sensitive and not is_tomes_pattern:
            tokens = pattern.split(" ")
            pattern = ["(?i)" + token + "(?-i)" for token in tokens]
            pattern = " ".join(pattern)
        elif not case_sensitive and is_tomes_pattern:
            msg = "Ignoring case insensitivity instruction for TOMES pattern in row {}.\
                    ".format(row_number)
            self.logger.warning(msg)

        # if @is_tomes_pattern is False, append @pattern to output.
        if not is_tomes_pattern:
            manifestations.append(pattern)

        return sorted(manifestations)

        
    def _get_rows(self, xlsx_file):
        """ Gets iterable version of row data for worksheet @self.entity_worksheet for a 
        given workbook, @xlsx_file.
        
        Args:
            - xlsx_file (str): The path to the Excel file to load.

        Returns:
            openpyxl.worksheet.worksheet.Worksheet: The return value.

        Raises:
            - KeyError: If @self.entity_worksheet is not present in @xlsx_file.
        """

        # load workbook.
        workbook = load_workbook(xlsx_file, read_only=False, data_only=True)
        
        # verify that required worksheet exists.
        try:
            entity_rows = workbook[self.entity_worksheet].iter_rows()
        except KeyError as err:
            self.logger.error(err)
            self.logger.warning("Missing required worksheet '{}' in workbook '{}'.".format(
                    self.entity_worksheet, xlsx_file))
            raise err

        return entity_rows


    def get_entities(self, xlsx_file):
        """ Gets rows in @xlsx_file and returns a generator. Each item in the generator is
        a dictionary in which the field names are keys and the row data are values.

        Args:
            - xlsx_file (str): The path to the Excel file to load.
        
        Returns:
            generator: The return value.
        
        Raises:
            self.SchemaError: If the header is invalid.
        """
 
        self.logger.info("Loading workbook: {}".format(xlsx_file))

        # report on total rows.
        total_rows = sum(1 for row in self._get_rows(xlsx_file))
        self.logger.info("Found {} rows.".format(total_rows))
        
        # get row data and modified checksum.
        entity_rows = self._get_rows(xlsx_file)
        hash_prefix = self._get_hash_prefix(xlsx_file)

        # get header.
        row = next(entity_rows)
        header = [cell.value for cell in row]
        header= tuple(header)

        # if header is invalid, return empty generator.
        if not self._validate_header(header):
            msg = "Invalid header row: {}.".format(header)
            raise self.SchemaError(msg)

        # create generator for each row.
        def entities():
            
            # start row numbering at 2 because the header row is the first.
            row_number = 2
            
            # yield a dict for each non-header row.
            header_range = range(0,len(header))
            for row in entity_rows:

                # get row values.
                row = [cell.value for cell in row]
                row = [cell.strip() if isinstance(cell, str) else cell for cell in row]
                row = [(header[i], row[i]) for i in header_range]
                row = dict(row)

                # run row validator.
                row_valid = self._validate_row(row, row_number)
                row_number += 1
                if not row_valid:
                    self.logger.warning("Skipping row {}; row is invalid.".format(
                        row_number - 1))
                    continue
                
                # alter data as needed and create dict for row.
                row["identifier"] = hash_prefix + row["identifier"]
                manifestations = self.get_manifestations(row["pattern"], row["case_sensitive"], 
                        row_number)
                row["manifestations"] = ["".join(m) for m in manifestations]
                
                # yield dict.
                yield(row)

        return entities()


    class SchemaError(Exception):
        """ A custom error class for invalid headers in TOMES Entity Dictionaries. """
        pass


if __name__ == "__main__":
    pass

