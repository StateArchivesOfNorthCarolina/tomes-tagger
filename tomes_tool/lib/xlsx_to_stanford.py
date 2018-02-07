#!/usr/bin/env python3

"""
This module contains a class for converting a TOMES Excel 2007+ (.xlsx) entity dictionary 
file to a Python list or a tab-delimited file containing NER mapping rules for Stanford
CoreNLP.

Todo:
    * Header validator needs to raise an error, NOT the tsv method.
    * Remove "if main: stuff at the end when you are ready.
    * Flesh out logging.
    * Gotta pre-count rows now that using generator.
        - Create a method to do this? Yes.
    * Gotta handle lists being returned from get_pattern().
        - Think this is OK now.
    * After you re-write any code, check your examples and docstrings.
"""

# import modules.
import codecs
import hashlib
import itertools
import logging
import os
from openpyxl import load_workbook


class XLSXToStanford():
    """ A class for converting a TOMES Excel 2007+ (.xlsx) entity dictionary file to a Python
    list or a tab-delimited file containing NER mapping rules for Stanford CoreNLP.
    
    Example:
        >>> x2s = XLSXToStanford()
        >>> entities = x2s.get_entities("entities.xlsx")
        >>> next(entities) # mapped dictionary for first data row in Excel file.
        >>> x2s.write_stanford("entities.xlsx", "entities.txt") # creates mapping file.
    """


    def __init__(self, entity_worksheet="Entities", charset="utf-8"):
        """ Sets instance attributes.
        
        Args:
            - entity_worksheet (str): The name of the worksheet to read from a given 
            workbook, i.e. an Excel file.
            - charset (str): Encoding used when writing to file.
        """
        
        # set logger; suppress logging by default. 
        self.logger = logging.getLogger(__name__)
        self.logger.addHandler(logging.NullHandler())

        # set attributes.
        self.entity_worksheet = entity_worksheet
        self.charset = charset
        self.required_headers = {"identifier":str, "pattern":str, "description":str,
                "case_sensitive":bool, "label":str, "authority":str}


    def _get_hash_prefix(self, xlsx_file):
        """ ??? this will be prepended to @identifier, below, in order to know the version
        of the source Excel file used. (first 6 characters should still be unique enough)
        
        Args:
            - ???
            
        Returns:
            str: ???
        """

        # get checksum of @xlsx_file; truncate it.
        checksum = hashlib.sha256()
        with open(xlsx_file, "rb") as xf:
            xf_bytes = xf.read()
        checksum.update(xf_bytes)
        hash_prefix = checksum.hexdigest()[:6] + "#"

        return hash_prefix


    def _validate_header(self, header_row):
        """ Determines if @header_row is equal to self.required_headers.
        
        Args:
            - header_row (tuple): The row values for the presumed first row of data.

        Returns:
            bool: The return value.
            True if the header is valid, otherwise False.
        """

        self.logger.info("Validating header row: {}".format(header_row))
        
        # ???
        is_valid = True
        
        # if header is 100 percent valid, return @is_valid.
        if header_row == self.required_headers.keys():
            self.logger.info("Header row is valid.")
            return is_valid

        # find any extra header fields.
        extra_headers = [header for header in header_row if header not in 
                self.required_headers]
        if len(extra_headers) != 0:
            self.logger.warning("Found extra headers: {}".format(extra_headers))

        # find any missing header fields.
        missing_headers = [header for header in self.required_headers if header not in
                header_row]
        if len(missing_headers) != 0:
            self.logger.error("??? no valid ")
            self.logger.warning("Missing headers: {}".format(missing_headers))
            is_valid = False
  
        return is_valid


    def _validate_row(self, row, line):
        """ ??? 
        
        Args:
            - row (dict): ???
            
        Returns:
            bool ???
        """

        # ???
        is_valid = True

        # ???
        tests = []
        for k,v in row.items():
            test = isinstance(v, self.required_headers[k])
            if not test:
                self.logger.warning("??? Field {} in row # {} not valid ...".format(k,
                    line))
            tests.append(test)

        # ???
        if False in tests:
            is_valid = False
  
        return is_valid


    def _check_tomes_pattern(self, pattern):
        """ """
        
        # ???
        is_tomes_pattern = False
        parts = pattern.split("tomes_pattern:")
        
        # ???
        parts_len = len(parts)
        if parts_len == 1:
            tomes_pattern_check = is_tomes_pattern, pattern
            return tomes_pattern_check
        elif parts_len > 2:
            self.logger.warning("??? Expected 2, got {} ... falling back to ...".format(
                parts_len))
            tomes_pattern_check = is_tomes_pattern, pattern
            return tomes_pattern_check            
        
        # ???
        pattern = parts[1]
        is_tomes_pattern = True
        self.logger.info("??? it's a tomes pattern ...")

            
        # ???
        if pattern[-1] != ",":
            pattern += ","
        
        # ???
        try:
            pattern = eval(pattern)
        except (NameError, SyntaxError) as err:
            self.logger.error(err)
            self.logger.warning("??? Invalid syntax ... falling back to ...")
        try:
            patterns = [i for i in itertools.product(*pattern)]
            patterns.reverse()
        except TypeError as err:
            self.logger.error(err)
            self.logger.warning("??? Check syntax. falling back to ...")

        return is_tomes_pattern, patterns


    def _get_patterns(self, pattern, case_sensitive):
        """ Returns manifestations @pattern without excess whitespace. If @case_sensitive is
        True, also alters @pattern to include case-insensitive regex markup.
        
        Args:
            - pattern (str): The "pattern" field value for a given row.
            - case_sensitive (bool): The "case_sensitive" field value for a given row.

        Returns:
            list: The return value.
            The altered versions of @pattern.
        """

        # remove excess whitespace from @pattern.
        _pattern = " ".join(pattern.strip().split())
        if pattern != _pattern:
            self.logger.warning("Removing excess whitespace in: {}".format(pattern))
            pattern = _pattern
        
        # ???
        is_tomes_pattern, patterns = self._check_tomes_pattern(pattern)

        # if specified, alter @pattern to ignore case provided @is_tomes_pattern is False.
        if not case_sensitive and not is_tomes_pattern:
            tokens = pattern.split(" ")
            pattern = ["(?i)" + token + "(?-i)" for token in tokens]
            pattern = " ".join(pattern)
        elif not case_sensitive and is_tomes_pattern:
            self.logger.warning("??? Ignoring case instruction because ...")
        
        # ???
        if not is_tomes_pattern:
            patterns = [pattern]
            
        return patterns

        
    def _get_rows(self, xlsx_file):
        """ Gets iterable version of row data for worksheet (self.entity_worksheet) for a 
        given workbook (@xlsx_file).
        
        Args:
            - xlsx_file (str): The path to the Excel file to load.

        Returns:
            openpyxl.worksheet.worksheet.Worksheet: The return value.
            ??? It's a genererator now.

        Raises:
            - ???
        """

        # load workbook.
        workbook = load_workbook(xlsx_file, read_only=False, data_only=True)
        
        # verify that required worksheet exists.
        try:
            entity_rows = workbook[self.entity_worksheet].iter_rows()
        except KeyError as err:
            self.logger.error(err)
            self.logger.warning("Missing required worksheet '{}' in workbook: {}".format(
                    self.entity_worksheet, xlsx_file))
            raise err

        return entity_rows


    def get_entities(self, xlsx_file):
        """ ??? Adds key "manifestations" ...

        Args:
            - ???
        
        Returns:
            list: ???

        Raises:
            - ???     
        """

        # load workbook; get row data and modified checksum.
        self.logger.info("Loading workbook: {}".format(xlsx_file))
        entity_rows = self._get_rows(xlsx_file)
        hash_prefix = self._get_hash_prefix(xlsx_file)

        # ???
        def entities():
        
            # get header.
            row = next(entity_rows)
            header = [cell.value for cell in row]
            header= tuple(header)

            # validate header.
            if not self._validate_header(header):
                raise Exception # ??? TODO what kind?
            
            # ???
            line_number = 1
            header_range = range(0,len(header))
            for row in entity_rows:
                
                # ???
                row = [cell.value for cell in row]
                row = [(header[i], row[i]) for i in header_range]
                row = dict(row)

                # ???
                self._validate_row(row, line_number)
                
                row["identifier"] = hash_prefix + row["identifier"]
                manifestations = self._get_patterns(row["pattern"], row["case_sensitive"])
                row["manifestations"] = ["".join(m) for m in manifestations]
                
                # ???
                line_number += 1
                yield(row)

        return entities()


    def write_stanford(self, xlsx_file, stanford_file):
        """ Converts @xlsx_file to a CoreNLP mapping file (@stanford_file).
        
        Args:
            - xlsx_file (str): The path to the Excel file to convert.
            - stanford_file (str): The output path for the converted file.

        Returns:
            None

        Raises:
            FileExistsError: If @stanford_file already exists.
        """
        
        # raise error if @stanford_file already exists.
        if os.path.isfile(stanford_file):
            err = "Destination file '{}' already exists.".format(stanford_file)
            self.logger.error(err)
            raise FileExistsError(err)

        # load workbook; get row data and modified checksum.
        self.logger.info("Loading workbook: {}".format(xlsx_file))
        entity_rows = self.get_entities(xlsx_file)
        hash_prefix = self._get_hash_prefix(xlsx_file)
       
        # open @stanford_file for writing.
        tsv = codecs.open(stanford_file, "w", encoding=self.charset)

        # iterate through rows; write data to @stanford_file.
        #???self.logger.info("Writing to mapping file: {}".format(stanford_file))
        row_count = 1
        total_rows = sum(1 for d in self.get_entities(xlsx_file))
        for row in entity_rows:
            
            # get cell data.
            tag = row["identifier"], row["authority"], row["label"]
            tag = "::".join(tag)
            manifestations = row["manifestations"]

            # ??? write row to file and avoid final linebreak (otherwise CoreNLP will crash).
            for manifestation in manifestations:
                tsv.write("\t".join([manifestation,tag]))
                if row_count != total_rows:
                    tsv.write("\n")
            row_count += 1

        return


if __name__ == "__main__":
    #pass
    import logging
    logging.basicConfig(level="DEBUG")
    x2s = XLSXToStanford()
    #x2s.write_stanford("../../NLP/TOMES_Entity_Dictionary.xlsx", "mapping.txt")
    #entities = x2s.get_entities("../../NLP/TOMES_Entity_Dictionary.xlsx")
    entities = x2s.get_entities("../../NLP/foo.xlsx")
    #p = x2s._get_patterns("tomes_pattern:{'[A|a]'}, {'B'}", True)
    #p = x2s._get_patterns("tomes_pattern:{'A', 'B'}, {'-',' '}, {'[0-9]{1,2}', '000'}", True)
    #print(p)
    for i in entities: print(i)
