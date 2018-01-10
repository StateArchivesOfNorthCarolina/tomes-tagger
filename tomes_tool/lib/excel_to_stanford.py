#!/usr/bin/env python3

"""
This module contains a class for converting an Excel 2007+ (.xlsx) file to a tab-delimited
file containing NER mapping rules for Stanford CoreNLP.

Todo:
    * Should you validate the data type for each row, i.e. "_validate_row()"?
        - No: Just trust that data is OK unless we start to see data entry errors. :-]
"""

# import modules.
import codecs
import glob
import logging
import os
import shutil
from openpyxl import load_workbook


class XLSXToStanford():
    """ A class for converting an Excel 2007+ (.xlsx) file to a tab-delimited file containing
    NER mapping rules for Stanford CoreNLP. 
    
    Example:
        >>> x2s = XLSXToStanford()
        >>> x2s.convert_file("foo.xlsx") # outputs "foo__mapping.txt".
        >>> xs2.convert_folder("in", "out") # converts "in/*.xlsx" to "out/*.txt".
    """


    def __init__(self, entity_worksheet="Entities", charset="utf-8"):
        """ Sets instance attributes.
        
        Args:
            - entity_worksheet (str): The name of the worksheet to read from a given workbook,
            i.e. an Excel file.
            - charset (str): Encoding used when writing to file.
        """
        
        # set logger; suppress logging by default. 
        self.logger = logging.getLogger(__name__)
        self.logger.addHandler(logging.NullHandler())

        # set attributes.
        self.entity_worksheet = entity_worksheet
        self.charset = charset
        self.required_headers = ("pattern", "description", "case_sensitive", "label", 
                "authority")


    def _get_pattern(self, pattern, case_sensitive):
        """ Returns @pattern without excess whitespace. If @case_sensitive is True, also 
        alters @pattern to include case-insensitive regex markup.
        
        Args:
            - pattern (str): The "pattern" field value for a given row.
            - case_sensitive (bool): The "case_sensitive" field value for a given row.

        Returns:
            str: The return value.
            The altered version of @pattern.
        """

        # remove excess whitespace from @pattern.
        _pattern = " ".join(pattern.strip().split())
        if pattern != _pattern:
            self.logger.warning("Removing excess whitespace in: {}".format(pattern))
            pattern = _pattern
        
        # if specified, alter @pattern to be case-insensitive.
        if not case_sensitive:
            pattern = "(?i)" + pattern.replace(" ", " (?i)")
        
        return pattern


    def _validate_header_row(self, header_row):
        """ Determines if @header_row is equal to self.required_headers.
        
        Args:
            - header_row (tuple): The row values for the presumed first row of data.

        Returns:
            bool: The return value.
            True if the header is valid, otherwise False.
        """
        
        self.logger.debug("Validating header row: {}".format(header_row))
        
        # if header is value, return True.
        if header_row == self.required_headers:
            self.logger.info("Header row is valid.")
            return True
        
        # otherwise, find and log errors.
        self.logger.error("Header row is not valid.")
        
        # find missing header fields.
        missing_headers = [header for header in self.required_headers if header not in
                header_row]
        if len(missing_headers) != 0:
            self.logger.warning("Missing headers: {}".format(missing_headers))

        # find extra header fields.
        extra_headers = [header for header in header_row if header not in 
                self.required_headers]
        if len(extra_headers) != 0:
            self.logger.warning("Found extra headers: {}".format(extra_headers))
  
        return False

        
    def _get_entity_rows(self, xlsx_file):
        """ Gets iterable version of row data for worksheet (self.entity_worksheet) for a 
        given workbook (@xlsx_file).
        
        Args:
            - xlsx_file (str): The path to the Excel file to load.

        Returns:
            openpyxl.worksheet.worksheet.Worksheet: The return value.    
        """

        # load workbook.
        workbook = load_workbook(xlsx_file, read_only=False, data_only=True)
        
        # verify that required worksheet exists.
        try:
            entity_rows = workbook[self.entity_worksheet]
        except KeyError:
            self.logger.error("Missing required worksheet '{}' in workbook: {}".format(
                    self.entity_worksheet, xlsx_file))
            raise

        return entity_rows


    def _unlink_outfile(self, output_file):
        """ Deletes @output_file if it already exists.
        
        Args:
            - output_file (str): The file to delete if it exists.
            
        Returns:
            None
        """

        if os.path.isfile(output_file):
            self.logger.info("Removing existing output file: {}".format(output_file))
            os.remove(output_file)
        
        return


    def convert_file(self, xlsx_file, tsv_file):
        """ Converts @xlsx_file to a CoreNLP mapping file (@tsv_file).
        
        Args:
            - xlsx_file (str): The path to the Excel file to convert.
            - tsv_file (str): The output path for the converted file.

        Returns:
            None
        """
            
        # load workbook; get row data.
        self.logger.info("Loading workbook: {}".format(xlsx_file))
        entity_rows = self._get_entity_rows(xlsx_file)
       
        # if a file named @tsv_file exists, delete it.
        self._unlink_outfile(tsv_file)

        # open @tsv_file for writing.
        tsv = codecs.open(tsv_file, "w", encoding=self.charset)

        # iterate through rows; write data to @tsv_file.
        i = 0
        for row in entity_rows.values:

            # validate header row.
            if i == 0:
                if not self._validate_header_row(row):
                    tsv.close()
                    self._unlink_file(tsv_file)
                    err = "Bad header in workbook: {}".format(xlsx_file)
                    raise Exception(err)
                else:
                    self.logger.info("Writing to mapping file: {}".format(tsv_file))
            
            # get cell data.
            pattern, description, case_sensitive, label, authority = row
            pattern = self._get_pattern(pattern, case_sensitive)
            label = authority + "/" + label

            # write row to file and avoid final linebreak (otherwise CoreNLP will crash).
            if i != 0:
                tsv.write("\t".join([pattern,label]))
                if i != entity_rows.max_row - 1:
                    tsv.write("\n")

            i += 1

        return


    def convert_folder(self, input_dir, output_dir):
        """ Globs all Excel files (non-recursive) in @input_dir and converts them with 
        self.convert_file(). Resultant files are placed in @output_dir.
        
        Args:
            - input_dir (str): The source directory containing Excel files to convert.
            - output_dir (str): The destination directory in which to placed the converted 
            files.
            
        Returns:
            None
        """

        # get absolute path to folders.
        input_dir = os.path.abspath(input_dir)
        output_dir = os.path.abspath(output_dir)

        # if directories match, log error and exit.
        if input_dir == output_dir:
            self.logger.error("Source and destination folder are identical. Aborting.")
            return

        # if a folder does not exist, raise an error.
        for _dir in [input_dir, output_dir]:
            if not os.path.isdir(_dir):
                err = "Non-existant directory: {}.".format(_dir)
                raise Exception(err)

        # glob mapping files; avoid temporary Excel files.
        xlsx_files = glob.glob(input_dir + "/[!~]*.xlsx")

        # convert Excel files and save to @output_dir.
        for xlsx_file in xlsx_files:
            out_file = xlsx_file.replace(input_dir, output_dir)
            out_file = out_file.replace(".xlsx", ".txt")
            self._unlink_outfile(out_file)
            self.logger.info("Converting '{}' to '{}'.".format(xlsx_file, out_file))
            self.convert_file(xlsx_file, out_file)

        return


if __name__ == "__main__":
    pass

